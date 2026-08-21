# -*- coding: utf-8 -*-
"""Tools pembuat file Shopee Mass Upload.

Perintah:
    python tools/shopee_mass_upload.py impor <file>  -> ubah ekspor sheet SKU jadi data/sku.csv
    python tools/shopee_mass_upload.py deteksi <folder> -> hitung foto di 1 folder, tanpa mengubah apa pun
    python tools/shopee_mass_upload.py unggah <folder> -> proses 1 folder produk:
                                                deteksi PNG, salin, push, simpan URL ke database
    python tools/shopee_mass_upload.py foto    -> salin & rename SEMUA foto dari Drive
    python tools/shopee_mass_upload.py url     -> daftar "file lokal -> URL" ke data/url_foto.csv
    python tools/shopee_mass_upload.py cek     -> laporan kelengkapan data & foto
    python tools/shopee_mass_upload.py build   -> hasilkan file Excel siap upload
    python tools/shopee_mass_upload.py semua   -> foto + url + cek + build

Input : data/sku.csv  (kolom: jenis, seri, varian, sku)  + tools/config.json
Output: output/*.xlsx

Kolom template dicari lewat kode field di baris 1 (mis. "ps_price|1|1"), bukan
lewat posisi kolom, supaya tetap jalan kalau Shopee mengubah susunan template
atau kalau dipakai untuk kategori lain.
"""
import argparse, csv, json, os, re, shutil, sys, time
from collections import OrderedDict

# openpyxl menolak file Shopee karena atribut activePane tidak baku -> longgarkan dulu
import openpyxl.worksheet.views as _views
import openpyxl.descriptors as _desc
_views.Pane.activePane = _desc.Typed(expected_type=str, allow_none=True)
import openpyxl

AKAR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG = os.path.join(AKAR, 'tools', 'config.json')
SKU_CSV = os.path.join(AKAR, 'data', 'sku.csv')
URL_CSV = os.path.join(AKAR, 'data', 'url_foto.csv')
DIR_FOTO = os.path.join(AKAR, 'foto-upload')
DIR_OUT = os.path.join(AKAR, 'output')
MANIFEST = os.path.join(DIR_FOTO, '_manifest.json')
DB_PATH = os.path.join(AKAR, 'data', 'foto.db')
LOKAL = os.path.join(AKAR, 'data', 'lokal.json')

# Saringan uji coba. Kalau salah satu diisi, tools hanya memproses yang cocok dan
# hasilnya ditulis ke folder terpisah (output/uji, data/uji) supaya berkas asli aman.
SARING = {'toko': None, 'jenis': None, 'seri': None}


def menyaring():
    return any(SARING.values())


def cocok(nilai, kunci):
    """Cocok kalau saringan kosong, atau teksnya terkandung (tidak peka huruf besar)."""
    pola = SARING.get(kunci)
    return not pola or pola.lower() in str(nilai).lower()


def dir_keluaran():
    return os.path.join(DIR_OUT, 'uji') if menyaring() else DIR_OUT


def path_url():
    if menyaring():
        return os.path.join(AKAR, 'data', 'uji', 'url_foto.csv')
    return URL_CSV


EKSTENSI = ('.png', '.jpg', '.jpeg')
BATAS_FOTO = 2 * 1024 * 1024        # batas ukuran foto Shopee
MAKS_JUDUL, MIN_DESK, MAKS_DESK = 255, 20, 3000
MAKS_NAMA_VARIASI, MAKS_VARIAN = 14, 20


def _timpa(dasar, atas):
    """Gabungkan dua dict bersarang; nilai di `atas` menang."""
    for k, v in (atas or {}).items():
        if isinstance(v, dict) and isinstance(dasar.get(k), dict):
            _timpa(dasar[k], v)
        elif v is not None:
            dasar[k] = v
    return dasar


def baca_config():
    """Pengaturan bersama dari tools/config.json, ditimpa data/lokal.json.

    config.json ikut diperbarui dari GitHub, jadi isinya hal yang sama untuk
    semua komputer: harga, deskripsi, judul, kategori. Hal yang berbeda tiap
    komputer — terutama letak folder foto di Drive — disimpan di data/lokal.json
    yang tidak ikut git, sehingga tidak tertimpa saat memperbarui.
    """
    with open(CONFIG, encoding='utf-8') as f:
        cfg = json.load(f)
    if os.path.exists(LOKAL):
        try:
            with open(LOKAL, encoding='utf-8') as f:
                _timpa(cfg, json.load(f))
        except ValueError:
            print('[config] data/lokal.json rusak, diabaikan')
    return cfg


def baca_lokal():
    if os.path.exists(LOKAL):
        try:
            with open(LOKAL, encoding='utf-8') as f:
                return json.load(f)
        except ValueError:
            pass
    return {}


def tulis_lokal(isi):
    os.makedirs(os.path.dirname(LOKAL), exist_ok=True)
    with open(LOKAL, 'w', encoding='utf-8') as f:
        json.dump(isi, f, ensure_ascii=False, indent=2)


def baca_sku():
    """Baca data/sku.csv -> {jenis: OrderedDict{seri: [ {varian, sku}, ... ]}}

    SKU yang cocok dengan pola di config.json -> abaikan_sku dilewati. Bawaannya
    varian CUSTOM, yang diupload manual dan tidak ikut mass upload.
    """
    if not os.path.exists(SKU_CSV):
        sys.exit('File input tidak ada: {}\nBuat dulu dengan kolom: jenis,seri,varian,sku'.format(SKU_CSV))
    try:
        pola = baca_config().get('abaikan_sku') or []
    except Exception:
        pola = []
    dilewati = 0
    data = OrderedDict()
    with open(SKU_CSV, encoding='utf-8-sig', newline='') as f:
        for i, baris in enumerate(csv.DictReader(f), start=2):
            baris = {(k or '').strip().lower(): (v or '').strip() for k, v in baris.items()}
            if not baris.get('sku'):
                continue
            for wajib in ('jenis', 'seri'):
                if not baris.get(wajib):
                    sys.exit('sku.csv baris {}: kolom "{}" kosong'.format(i, wajib))
            # kolom "varian" boleh tidak ada; nama varian memakai SKU-nya
            baris.setdefault('varian', '') or baris.update(varian=baris['sku'])
            if any(re.search(x, baris['sku'], re.I) for x in pola):
                dilewati += 1
                continue
            kode = baris.get('kode_seri') or kode_seri(baris['seri'])
            if not cocok(baris['jenis'], 'jenis'):
                continue
            if not (cocok(baris['seri'], 'seri') or cocok(kode, 'seri')):
                continue
            data.setdefault(baris['jenis'].upper(), OrderedDict()) \
                .setdefault(baris['seri'], []).append(
                    {'varian': baris['varian'], 'sku': baris['sku'], 'kode_seri': kode})
    if dilewati and not baca_sku._diam:
        print('[input] {} SKU dilewati sesuai abaikan_sku di config.json'.format(dilewati))
        baca_sku._diam = True
    if not data:
        sys.exit('Tidak ada SKU yang cocok dengan saringan: {}'.format(
            {k: v for k, v in SARING.items() if v}))
    return data


baca_sku._diam = False


DIR_TEMPLATE = os.path.join(AKAR, 'template')


def kategori_template(path):
    """Daftar kategori yang didukung sebuah berkas template (dari HiddenCatProps)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if 'HiddenCatProps' not in wb.sheetnames:
            return []
        hp = wb['HiddenCatProps']
        return [hp.cell(r, 1).value for r in range(1, hp.max_row + 1) if hp.cell(r, 1).value]
    finally:
        wb.close()


def info_template(cfg):
    """Keterangan tiap template yang terpasang: umur berkas, kategori, dan
    apakah kategori yang dipakai di config benar-benar ada di dalamnya."""
    dipakai = {}
    for jenis, j in cfg['jenis'].items():
        dipakai.setdefault(j['template'], []).append((jenis, j['kategori']))
    hasil = []
    for kunci, path in cfg['template'].items():
        penuh = os.path.join(AKAR, path)
        ada = os.path.exists(penuh)
        daftar = kategori_template(penuh) if ada else []
        kurang = [k for _, k in dipakai.get(kunci, []) if k not in daftar]
        hasil.append({
            'kunci': kunci, 'path': path, 'ada': ada,
            'kategori': len(daftar),
            'umur_hari': round((time.time() - os.path.getmtime(penuh)) / 86400, 1) if ada else None,
            'dipakai': [x[0] for x in dipakai.get(kunci, [])],
            'kategori_hilang': kurang,
        })
    return hasil


def pasang_template(cfg, berkas):
    """Pasang berkas template Shopee yang baru diunduh.

    Tujuannya ditentukan dari isi berkas itu sendiri: kategori mana yang
    didukungnya dicocokkan dengan kategori yang dipakai di config, jadi tidak
    perlu memilih-milih secara manual.
    """
    if not os.path.exists(berkas):
        sys.exit('Berkas tidak ditemukan: {}'.format(berkas))
    daftar = kategori_template(berkas)
    if not daftar:
        sys.exit('Berkas itu tidak terlihat seperti template mass upload Shopee '
                 '(sheet HiddenCatProps tidak ada).')

    cocok = set()
    for jenis, j in cfg['jenis'].items():
        if j['kategori'] in daftar:
            cocok.add(j['template'])
    if not cocok:
        print('[template] {} kategori terbaca, tapi tidak satu pun cocok dengan '
              'kategori di config.json:'.format(len(daftar)))
        for jenis, j in cfg['jenis'].items():
            print('   {:<12} butuh: {}'.format(jenis, j['kategori']))
        print('   Contoh kategori di berkas itu:')
        for k in daftar[:5]:
            print('     - {}'.format(k))
        sys.exit('Unduh template untuk cabang kategori yang sesuai.')
    if len(cocok) > 1:
        sys.exit('Berkas itu cocok untuk lebih dari satu template: {}'.format(sorted(cocok)))

    kunci = cocok.pop()
    tujuan = os.path.join(AKAR, cfg['template'][kunci])
    os.makedirs(os.path.dirname(tujuan), exist_ok=True)
    if os.path.exists(tujuan):
        cadangan = tujuan + '.bak'
        shutil.copy2(tujuan, cadangan)
        print('[template] versi lama dicadangkan ke {}'.format(os.path.basename(cadangan)))
    shutil.copy2(berkas, tujuan)
    print('[template] "{}" dipasang sebagai template {}'.format(
        os.path.basename(berkas), kunci))
    print('[template] {} kategori didukung, dipakai oleh: {}'.format(
        len(daftar),
        ', '.join(j for j, x in cfg['jenis'].items() if x['template'] == kunci)))
    return kunci


def saring_lingkup(data, lingkup):
    """Batasi data SKU ke folder-folder yang dipilih pengguna.

    lingkup = [{'jenis': 'JIBBITZ', 'dari': 1, 'sampai': 50}, ...]
    Satu seri ikut terpilih kalau ada SKU-nya yang jatuh di salah satu rentang.
    Daftar kosong berarti tanpa batasan (semua data dipakai).
    """
    if not lingkup:
        return data
    hasil = OrderedDict()
    for jenis, seri_map in data.items():
        rentang = [(L['dari'], L['sampai']) for L in lingkup if L.get('jenis') == jenis]
        if not rentang:
            continue
        for seri, desain in seri_map.items():
            for d in desain:
                n = nomor_sku(d['sku'])
                if n and any(a <= n <= b for a, b in rentang):
                    hasil.setdefault(jenis, OrderedDict())[seri] = desain
                    break
    return hasil


def dir_jenis(cfg, jenis):
    """Folder sumber di komputer ini untuk satu jenis produk.

    Dipakai "path_drive" kalau diisi (boleh drive/komputer mana saja). Kalau
    kosong, dirakit dari foto.root + folder_drive seperti pengaturan lama.
    """
    j = cfg['jenis'][jenis]
    khusus = (j.get('path_drive') or '').strip()
    if khusus:
        return os.path.normpath(khusus)
    akar = (cfg.get('foto', {}).get('root') or '').strip()
    return os.path.normpath(os.path.join(akar, j.get('folder_drive') or jenis))


def kode_seri(nama_seri):
    """'... KPOP - CORTIS SERIES' -> 'CORTIS'"""
    ekor = nama_seri.split(' - ')[-1]
    return re.sub(r'\s*SERIES\s*$', '', ekor, flags=re.I).replace(' ', '').upper()


def nomor_sku(sku):
    """'JB-0000051' -> 51 ; 'JB-CUSTOM-01' -> None"""
    m = re.match(r'^[A-Z]+-(\d+)$', sku.upper())
    return int(m.group(1)) if m else None


# --------------------------------------------------------------------------- impor
JUDUL_KOLOM = {'sku': 'sku', 'varian': 'varian', 'nama produk': 'seri'}


def _baris_tabel(rows):
    """Dari daftar baris mentah, ambil dict {seri, varian, sku} memakai baris judul."""
    hasil, kolom = [], None
    for baris in rows:
        sel = [str(x).strip() if x is not None else '' for x in baris]
        rendah = [s.lower() for s in sel]
        if 'sku' in rendah and 'varian' in rendah and 'nama produk' in rendah:
            kolom = {JUDUL_KOLOM[s]: i for i, s in enumerate(rendah) if s in JUDUL_KOLOM}
            continue
        if not kolom:
            continue
        try:
            baris_data = {k: sel[i] for k, i in kolom.items()}
        except IndexError:
            continue
        if re.match(r'^[A-Z]{2}-[A-Z0-9-]+$', baris_data['sku'].upper()) and baris_data['varian']:
            hasil.append(baris_data)
    return hasil


POLA_SKU = re.compile(r'^[A-Z]{2}-[A-Z0-9][A-Z0-9-]*$', re.I)


def baca_tempelan(teks):
    """Baca data SKU yang ditempel langsung dari Google Sheet.

    Menerima hasil salin dari Sheets (dipisah tab) maupun CSV. Kalau ada baris
    judul yang memuat "SKU" dan "Nama Produk", kolomnya diambil dari situ.
    Kalau tidak ada, kolom SKU dikenali dari polanya (JB-0000001) dan kolom
    nama produk diambil dari kolom teks terpanjang.
    """
    baris = [b for b in teks.replace('\r', '').split('\n') if b.strip()]
    if not baris:
        return []
    pisah = '\t' if any('\t' in b for b in baris) else ','
    tabel = [[sel.strip().strip('"') for sel in b.split(pisah)] for b in baris]

    kolom, mulai = None, 0
    for i, sel in enumerate(tabel[:10]):
        rendah = [s.lower() for s in sel]
        if 'sku' in rendah and any(x in rendah for x in ('nama produk', 'seri')):
            kolom = {'sku': rendah.index('sku')}
            kolom['seri'] = rendah.index('nama produk') if 'nama produk' in rendah \
                else rendah.index('seri')
            if 'varian' in rendah:
                kolom['varian'] = rendah.index('varian')
            mulai = i + 1
            break

    if kolom is None:
        # tebak: kolom yang isinya berpola SKU, dan kolom teks terpanjang
        skor_sku, panjang = {}, {}
        for sel in tabel:
            for i, s in enumerate(sel):
                if POLA_SKU.match(s):
                    skor_sku[i] = skor_sku.get(i, 0) + 1
                panjang[i] = max(panjang.get(i, 0), len(s))
        if not skor_sku:
            return []
        i_sku = max(skor_sku, key=skor_sku.get)
        sisa = {i: p for i, p in panjang.items() if i != i_sku}
        kolom = {'sku': i_sku, 'seri': max(sisa, key=sisa.get) if sisa else i_sku}

    hasil = []
    for sel in tabel[mulai:]:
        try:
            sku = sel[kolom['sku']]
            seri = sel[kolom['seri']]
        except IndexError:
            continue
        if not POLA_SKU.match(sku) or not seri:
            continue
        hasil.append({'sku': sku.upper(), 'seri': seri,
                      'varian': sel[kolom['varian']] if kolom.get('varian') is not None
                      and len(sel) > kolom['varian'] else sku.upper()})
    return hasil


def tulis_sku(cfg, catatan, gabung=True):
    """Tulis daftar SKU ke data/sku.csv. `gabung` menambah ke data yang ada."""
    jenis_dari_prefix = {j['prefix_sku'].upper(): nama for nama, j in cfg['jenis'].items()}
    lama = OrderedDict()
    if gabung and os.path.exists(SKU_CSV):
        with open(SKU_CSV, encoding='utf-8-sig', newline='') as f:
            for r in csv.DictReader(f):
                if r.get('sku'):
                    lama[r['sku'].upper()] = [r.get('jenis', ''), r.get('seri', ''),
                                              r.get('varian') or r['sku'], r['sku'].upper()]

    baru = diperbarui = dilewati = 0
    for c in catatan:
        jenis = jenis_dari_prefix.get(c['sku'].split('-')[0].upper())
        if not jenis:
            dilewati += 1
            continue
        isi = [jenis, c['seri'], c['varian'], c['sku']]
        if c['sku'] in lama:
            if lama[c['sku']] != isi:
                diperbarui += 1
            lama[c['sku']] = isi
        else:
            lama[c['sku']] = isi
            baru += 1

    if not lama:
        return {'ok': False, 'pesan': 'Tidak ada SKU yang bisa dibaca.'}
    if os.path.exists(SKU_CSV):
        shutil.copy2(SKU_CSV, SKU_CSV + '.bak')
    os.makedirs(os.path.dirname(SKU_CSV), exist_ok=True)
    with open(SKU_CSV, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['jenis', 'seri', 'varian', 'sku'])
        w.writerows(sorted(lama.values(), key=lambda r: (r[0], r[3])))
    return {'ok': True, 'baru': baru, 'diperbarui': diperbarui,
            'dilewati': dilewati, 'total': len(lama)}


def perintah_impor(cfg, sumber):
    """Ubah ekspor sheet SKU (.xlsx / .csv) menjadi data/sku.csv."""
    if not os.path.exists(sumber):
        sys.exit('File tidak ditemukan: {}'.format(sumber))
    jenis_dari_prefix = {j['prefix_sku'].upper(): nama for nama, j in cfg['jenis'].items()}

    baris = []
    if sumber.lower().endswith(('.xlsx', '.xlsm')):
        wb = openpyxl.load_workbook(sumber, data_only=True)
        for ws in wb.worksheets:
            baris += _baris_tabel(ws.iter_rows(values_only=True))
        wb.close()
    else:
        with open(sumber, encoding='utf-8-sig', newline='') as f:
            baris += _baris_tabel(csv.reader(f))

    keluar, dilewati, terlihat = [], {}, set()
    for b in baris:
        sku = b['sku'].upper()
        if sku in terlihat:
            continue
        terlihat.add(sku)
        jenis = jenis_dari_prefix.get(sku.split('-')[0])
        if not jenis:
            dilewati[sku.split('-')[0]] = dilewati.get(sku.split('-')[0], 0) + 1
            continue
        keluar.append([jenis, b['seri'], b['varian'], sku])

    if not keluar:
        sys.exit('Tidak ada baris SKU yang terbaca. Pastikan file punya kolom '
                 '"Nama Produk", "Varian", dan "SKU".')

    keluar.sort(key=lambda r: (r[0], r[3]))
    if os.path.exists(SKU_CSV):
        shutil.copy2(SKU_CSV, SKU_CSV + '.bak')
        print('[impor] file lama dicadangkan ke sku.csv.bak')
    os.makedirs(os.path.dirname(SKU_CSV), exist_ok=True)
    with open(SKU_CSV, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['jenis', 'seri', 'varian', 'sku'])
        w.writerows(keluar)

    print('[impor] {} SKU ditulis ke {}'.format(len(keluar), SKU_CSV))
    per_jenis = {}
    for j, seri, _, _ in keluar:
        per_jenis.setdefault(j, set()).add(seri)
    for j in sorted(per_jenis):
        n = sum(1 for r in keluar if r[0] == j)
        print('   {:<14} {} SKU / {} seri'.format(j, n, len(per_jenis[j])))
    for pre, n in dilewati.items():
        print('   ! prefix SKU "{}" tidak ada di config.json ({} baris dilewati)'.format(pre, n))


# --------------------------------------------------------------------------- foto
def cari_folder_seri(dir_jenis, nomor):
    """Cari folder 'PRODUK 00051 - 00100' yang memuat nomor SKU tertentu."""
    if not os.path.isdir(dir_jenis):
        return None
    for nama in sorted(os.listdir(dir_jenis)):
        m = re.match(r'^PRODUK\s+0*(\d+)\s*-\s*0*(\d+)$', nama, re.I)
        if m and int(m.group(1)) <= nomor <= int(m.group(2)):
            return os.path.join(dir_jenis, nama)
    return None


def salin_muat(src, dst):
    """Salin foto; kalau > 2 MB dikecilkan sampai lolos batas Shopee."""
    if os.path.getsize(src) <= BATAS_FOTO:
        shutil.copy2(src, dst)
        return False
    from PIL import Image
    im = Image.open(src)
    if im.mode not in ('RGB', 'RGBA'):
        im = im.convert('RGBA')
    lebar = im.width
    for _ in range(8):
        im.resize((lebar, round(im.height * lebar / im.width)), Image.LANCZOS).save(dst, optimize=True)
        if os.path.getsize(dst) <= BATAS_FOTO:
            break
        lebar = int(lebar * 0.8)
    return True


def perintah_foto(cfg, data):
    """Salin foto dari Drive ke foto-upload/<toko>/<jenis>/ dengan nama rapi."""
    fcfg = cfg['foto']
    peta_utama = {n.lower(): i + 1 for i, n in enumerate(fcfg['nama_foto_utama'])}
    manifest, dikecilkan, dilewati, tak_ketemu = {}, [], [], []

    for jenis, seri_map in data.items():
        j = cfg['jenis'][jenis]
        dir_produk = dir_jenis(cfg, jenis)
        pre = j['prefix_sku']
        for seri, desain in seri_map.items():
            nomor = next((nomor_sku(d['sku']) for d in desain if nomor_sku(d['sku'])), None)
            folder = cari_folder_seri(dir_produk, nomor) if nomor else None
            dir_fp = os.path.join(folder, fcfg['subfolder']) if folder else None
            if not dir_fp or not os.path.isdir(dir_fp):
                tak_ketemu.append('{} / {}'.format(jenis, seri))
                continue
            kode = desain[0]['kode_seri']
            for sub in sorted(os.listdir(dir_fp)):
                dir_toko = os.path.join(dir_fp, sub)
                m = re.search(r'(\d)', sub)
                if not os.path.isdir(dir_toko) or not m:
                    continue
                toko = 'toko' + m.group(1)
                if not (cocok(toko, 'toko') or cocok(sub, 'toko')
                        or cocok(_nama_toko_dari(cfg, toko), 'toko')):
                    continue
                tujuan = os.path.join(DIR_FOTO, toko, j['slug'])
                os.makedirs(tujuan, exist_ok=True)
                for f in sorted(os.listdir(dir_toko)):
                    if not f.lower().endswith(EKSTENSI):
                        continue
                    if re.match(r'^' + pre + r'-\d+\.(png|jpe?g)$', f, re.I):
                        nama = f.upper().replace('.PNG', '.png').replace('.JPEG', '.jpeg').replace('.JPG', '.jpg')
                    elif f.lower() in peta_utama:
                        nama = '{}-{}-utama{}.png'.format(pre, kode, peta_utama[f.lower()])
                    else:
                        dilewati.append(os.path.join(j['folder_drive'], os.path.basename(folder), sub, f))
                        continue
                    if salin_muat(os.path.join(dir_toko, f), os.path.join(tujuan, nama)):
                        dikecilkan.append(nama)
                    manifest.setdefault(toko, {}).setdefault(j['slug'], {})[os.path.splitext(nama)[0]] = \
                        '{}/{}/{}'.format(toko, j['slug'], nama)

    os.makedirs(DIR_FOTO, exist_ok=True)
    with open(MANIFEST, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, ensure_ascii=False, indent=1)

    print('[foto] selesai')
    for toko in sorted(manifest):
        for slug in sorted(manifest[toko]):
            print('   {}/{}: {} file'.format(toko, slug, len(manifest[toko][slug])))
    print('   dikecilkan (>2MB): {} | dilewati (bukan foto produk): {}'.format(len(dikecilkan), len(dilewati)))
    for s in tak_ketemu:
        print('   ! folder foto belum ada: {}'.format(s))


# --------------------------------------------------------------------------- url
def baca_manifest():
    if not os.path.exists(MANIFEST):
        return {}
    with open(MANIFEST, encoding='utf-8') as f:
        return json.load(f)


def _nama_toko_dari(cfg, folder_toko):
    for t in cfg['toko']:
        if t['folder_foto'] == folder_toko:
            return t['nama']
    return folder_toko


def _kunci_toko(nama):
    """'TOKO 1' / 'Toko_1' / 'toko1' -> 'toko1'. None kalau tidak ada angkanya."""
    m = re.search(r'(\d+)', nama)
    return 'toko' + m.group(1) if m else None


def pindai_foto(cfg):
    """Pindai isi foto-upload/ dan kelompokkan per toko + jenis produk.

    Menelusuri semua sub-folder, jadi susunannya boleh:
        foto-upload/toko1/jibbitz/JB-0000001.png     (hasil perintah "foto")
        foto-upload/toko1/JB-0000001.png             (langsung di folder toko)
        foto-upload/TOKO 3/apa saja/PB-0000001.png   (sub-folder bebas)
    Jenis produk diambil dari nama sub-folder; kalau tidak cocok, ditebak dari
    prefix nama file (JB/PA/PB). Kembaliannya {toko: {slug: {kunci: path}}},
    dengan path relatif terhadap akar project (selalu diawali "foto-upload/")
    sehingga langsung cocok dengan base_url yang menunjuk akar repo.
    """
    hasil, tak_dikenal = {}, []
    if not os.path.isdir(DIR_FOTO):
        return hasil, tak_dikenal
    slug_dari_prefix = {j['prefix_sku'].upper(): j['slug'] for j in cfg['jenis'].values()}
    slug_sah = {j['slug'] for j in cfg['jenis'].values()}
    nama_jenis = {j['slug']: nama for nama, j in cfg['jenis'].items()}

    for folder in sorted(os.listdir(DIR_FOTO)):
        akar_toko = os.path.join(DIR_FOTO, folder)
        toko = _kunci_toko(folder)
        if not os.path.isdir(akar_toko) or folder.startswith(('.', '_')) or not toko:
            continue
        if not (cocok(toko, 'toko') or cocok(folder, 'toko')
                or cocok(_nama_toko_dari(cfg, toko), 'toko')):
            continue
        for dirpath, _, berkas in os.walk(akar_toko):
            if os.sep + '.' in dirpath:
                continue
            bagian = os.path.relpath(dirpath, akar_toko).split(os.sep)
            slug_folder = next((b for b in bagian if b in slug_sah), None)
            for f in sorted(berkas):
                if not f.lower().endswith(EKSTENSI):
                    continue
                kunci = os.path.splitext(f)[0].upper()
                slug = slug_folder or slug_dari_prefix.get(kunci.split('-')[0])
                rel = os.path.relpath(os.path.join(dirpath, f), AKAR).replace(os.sep, '/')
                if not slug:
                    tak_dikenal.append(rel)
                    continue
                if not (cocok(slug, 'jenis') or cocok(nama_jenis.get(slug, ''), 'jenis')):
                    continue
                hasil.setdefault(toko, {}).setdefault(slug, {})[kunci] = rel
    return hasil, tak_dikenal


JUDUL_URL = ['nama_toko', 'folder_toko', 'jenis', 'kunci', 'tipe',
             'file_lokal', 'url', 'ukuran_byte']


def perintah_url(cfg, data=None):
    """Pindai foto per toko, lalu tulis daftar 'file lokal -> URL'.

    Menghasilkan satu berkas gabungan (data/url_foto.csv) dan satu berkas per
    toko (data/url/url_foto - <Nama Toko>.csv) supaya mudah dipakai saat
    mengupload toko tertentu.
    """
    base = (cfg['foto'].get('base_url') or '').rstrip('/')
    if not base:
        sys.exit('foto.base_url di config.json masih kosong.\n'
                 'Isi dulu, contoh: "https://cdn.jsdelivr.net/gh/username/repo@main"')

    indeks, tak_dikenal = pindai_foto(cfg)
    if not indeks:
        sys.exit('Tidak ada foto di {}. Jalankan perintah "foto" dulu.'.format(DIR_FOTO))

    nama_toko = {t['folder_foto']: t['nama'] for t in cfg['toko']}
    jenis_slug = {j['slug']: nama for nama, j in cfg['jenis'].items()}

    # kalau menyaring per seri, hanya SKU pada seri itu (plus foto utamanya) yang diambil
    izin = None
    if SARING.get('seri') and data:
        izin = set()
        for seri_map in data.values():
            for desain in seri_map.values():
                izin.update(d['sku'].upper() for d in desain)
                izin.update('{}-UTAMA{}'.format(desain[0]['kode_seri'].upper(), n) for n in (1, 2, 3))

    def diizinkan(kunci):
        if izin is None:
            return True
        return kunci in izin or any(kunci.endswith(x) for x in izin if 'UTAMA' in x)

    per_toko, semua = OrderedDict(), []
    for toko in sorted(indeks):
        label = nama_toko.get(toko)
        if not label:
            print('   ! folder "{}" tidak terdaftar di config.json -> toko -> folder_foto'.format(toko))
            label = toko
        baris = []
        for slug in sorted(indeks[toko]):
            for kunci, rel in sorted(indeks[toko][slug].items()):
                if not diizinkan(kunci):
                    continue
                lokal = os.path.join(AKAR, rel.replace('/', os.sep))
                baris.append([label, toko, jenis_slug.get(slug, slug), kunci,
                              'utama' if '-UTAMA' in kunci.upper() else 'varian',
                              lokal, base + '/' + rel,
                              os.path.getsize(lokal) if os.path.exists(lokal) else ''])
        per_toko[label] = baris
        semua += baris

    def tulis(path, baris):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(JUDUL_URL)
            w.writerows(baris)

    berkas_utama = path_url()
    tulis(berkas_utama, semua)
    dir_toko = os.path.join(os.path.dirname(berkas_utama), 'url')
    for label, baris in per_toko.items():
        aman = re.sub(r'[\\/:*?"<>|]', '_', label)
        tulis(os.path.join(dir_toko, 'url_foto - {}.csv'.format(aman)), baris)

    print('[url] base: {}'.format(base))
    print('[url] {} URL -> {}'.format(len(semua), berkas_utama))
    for label, baris in per_toko.items():
        rinci = OrderedDict()
        for b in baris:
            rinci[b[2]] = rinci.get(b[2], 0) + 1
        utama = sum(1 for b in baris if b[4] == 'utama')
        print('   {:<15} {:>4} URL  ({} utama / {} varian)  {}'.format(
            label, len(baris), utama, len(baris) - utama,
            ' · '.join('{} {}'.format(v, k.lower()) for k, v in rinci.items())))
    print('[url] per toko -> {}'.format(dir_toko))

    # Foto hanya bisa diakses Shopee setelah ada di repo. Tunjukkan persis apa yang
    # perlu di-upload — berguna waktu uji coba sebagian.
    rel_semua = sorted({b[6][len(base) + 1:] for b in semua})
    folder = sorted({os.path.dirname(r) for r in rel_semua})
    print('[url] {} file di {} folder perlu ada di repo:'.format(len(rel_semua), len(folder)))
    for f in folder[:8]:
        n = sum(1 for r in rel_semua if os.path.dirname(r) == f)
        print('       {}/  ({} file)'.format(f, n))
    if len(folder) > 8:
        print('       ... dan {} folder lain'.format(len(folder) - 8))
    print('[url] perintah upload:')
    print('       cd "{}"'.format(AKAR))
    for f in folder[:4]:
        print('       git add -f "{}"'.format(f.replace('/', os.sep)))
    if len(folder) > 4:
        print('       (dan {} folder lainnya)'.format(len(folder) - 4))
    print('       git commit -m "foto uji coba" && git push')

    besar = [b for b in semua if b[7] and b[7] > BATAS_FOTO]
    if besar:
        print('   ! {} foto melebihi 2 MB, akan ditolak Shopee:'.format(len(besar)))
        for b in besar[:5]:
            print('       {} ({:.2f} MB)'.format(b[3], b[7] / 1024 ** 2))
    for rel in tak_dikenal[:5]:
        print('   ! jenis produk tidak dikenali, dilewati: {}'.format(rel))
    if len(tak_dikenal) > 5:
        print('   ! ... total {} file dilewati'.format(len(tak_dikenal)))


# --------------------------------------------------------------------------- excel
def peta_kolom(ws):
    """Kode field di baris 1 -> nomor kolom. 'ps_price|1|1' menjadi kunci 'ps_price'."""
    peta = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            peta.setdefault(str(v).split('|')[0].strip(), c)
    return peta


def atribut_wajib(wb, kategori):
    """Daftar id atribut yang MANDATORY untuk kategori ini (dari sheet HiddenCatProps)."""
    if 'HiddenCatProps' not in wb.sheetnames:
        return {}
    hp, tpl = wb['HiddenCatProps'], wb['Template']
    for r in range(1, hp.max_row + 1):
        if hp.cell(r, 1).value == kategori:
            hasil = {}
            for c in range(2, hp.max_column + 1):
                if str(hp.cell(r, c).value).upper() == 'MANDATORY':
                    kode = str(tpl.cell(1, c).value or '').split('|')[0]
                    if '.' in kode:
                        hasil[kode.split('.')[-1]] = tpl.cell(3, c).value
            return hasil
    return {}


def kunci_tambahan(punya, slug):
    """Kunci foto tambahan milik satu toko, urut, untuk satu jenis produk.

    Yang khusus jenis itu (TAMBAHAN-JIBBITZ-1, -2, ...) didahulukan. Kalau tidak
    ada, dipakai yang berlaku umum (TAMBAHAN-U1, -U2, ... dan "TAMBAHAN" lama).
    """
    khusus = sorted(k for k in punya if k.startswith('TAMBAHAN-' + slug.upper() + '-'))
    if khusus:
        return khusus
    return sorted(k for k in punya
                  if k == 'TAMBAHAN' or re.match(r'^TAMBAHAN-U\d+$', k))


def susun_listing(cfg, data, toko, jenis, manifest, dari_db=None):
    """Bangun daftar listing untuk satu toko + satu jenis produk."""
    j = cfg['jenis'][jenis]
    profil = cfg['profil'][toko['profil']]
    opsi = profil['judul'][jenis]
    teks = profil['deskripsi'][jenis].replace('{SPEC}', j['spec']).replace('{TOKO}', toko['nama'])
    base = (cfg['foto'].get('base_url') or '').rstrip('/')
    db_toko = (dari_db or {}).get(toko['folder_foto'], {})
    punya = (manifest or {}).get(toko['folder_foto'], {}).get(j['slug'], {})

    def url(kunci):
        # kunci selalu huruf besar, samakan supaya "utama1" ikut cocok
        kunci = kunci.upper()
        if db_toko:
            return db_toko.get(kunci)
        p = punya.get(kunci)
        return base + '/' + p if (base and p) else None

    hasil = []
    for i, (seri, desain) in enumerate(data[jenis].items()):
        depan, belakang = opsi[i % len(opsi)]
        kode = desain[0]['kode_seri']
        utama = [url('{}-{}-utama{}'.format(j['prefix_sku'], kode, n)) for n in (1, 2, 3)]
        # Foto per varian bersifat semua-atau-tidak-sama-sekali. Varian tanpa foto
        # sendiri (mis. CUSTOM) memakai foto utama sebagai pengganti.
        per_varian = [url(d['sku']) or utama[0] for d in desain]
        if not all(per_varian):
            per_varian = [None] * len(desain)
        # Foto tambahan milik toko (panduan ukuran) dipakai di semua listing.
        # Yang khusus satu jenis produk didahulukan daripada yang berlaku umum.
        tambahan = [db_toko[k] for k in kunci_tambahan(db_toko, j['slug'])]
        hasil.append({
            'jenis': jenis,
            'judul': '{} - {} - {}'.format(depan, seri, belakang),
            'deskripsi': teks,
            # SKU induk memakai SKU pertama produk itu, sesuai penomoran di sheet
            'sku_induk': desain[0]['sku'],
            'kode_induk': '{}-{}-{}'.format(toko['kode'], j['prefix_sku'], kode),
            'desain': desain, 'utama': utama, 'per_varian': per_varian,
            'tambahan': tambahan,
        })
    return hasil


def tulis_excel(cfg, path_template, path_out, listings):
    wb = openpyxl.load_workbook(path_template)
    ws = wb['Template']
    K = peta_kolom(ws)

    def isi(baris, kode, nilai):
        if nilai is not None and kode in K:
            ws.cell(baris, K[kode]).value = nilai

    kanal = [k for k in K if k.startswith('channel_id.')]
    r = 7
    for L in listings:
        j = cfg['jenis'][L['jenis']]
        harga = round(j['harga_paket'] / j['min_order'])
        for n, d in enumerate(L['desain']):
            if n == 0:
                isi(r, 'ps_product_name', L['judul'])
                isi(r, 'ps_product_description', L['deskripsi'])
                isi(r, 'ps_sku_parent_short', L['sku_induk'])
                isi(r, 'ps_item_cover_image', L['utama'][0])
                isi(r, 'ps_item_image_1', L['utama'][1])
                isi(r, 'ps_item_image_2', L['utama'][2])
                # foto tambahan mengisi Foto Produk 3 dan seterusnya
                # (pakai nama sendiri; "n" dipakai loop varian di luar)
                for slot, tautan in enumerate(L['tambahan'], start=3):
                    if slot > 8:
                        break
                    isi(r, 'ps_item_image_{}'.format(slot), tautan)
            isi(r, 'ps_category', j['kategori'])
            isi(r, 'ps_minimum_purchase_quantity', j['min_order'])
            isi(r, 'et_title_variation_integration_no', L['kode_induk'])
            isi(r, 'et_title_variation_1', 'Desain')
            # nama varian memakai kode SKU-nya langsung, bukan "Desain 01"
            isi(r, 'et_title_option_for_variation_1', d['sku'])
            isi(r, 'et_title_image_per_variation', L['per_varian'][n])
            isi(r, 'ps_price', harga)
            isi(r, 'ps_stock', j['stok'])
            isi(r, 'ps_sku_short', d['sku'])
            isi(r, 'ps_weight', j['berat_gram'])
            for k in kanal:
                isi(r, k, 'Aktif')
            for id_attr, nilai in j['atribut'].items():
                isi(r, 'ps_product_global_attribute.' + id_attr, nilai)
            r += 1

    os.makedirs(os.path.dirname(path_out), exist_ok=True)
    wb.save(path_out)
    return r - 7


def periksa(cfg, listings, wajib):
    """Validasi terhadap batasan Shopee. Kembalikan daftar peringatan."""
    pesan = []
    for L in listings:
        n = len(L['desain'])
        if not 5 <= len(L['judul']) <= MAKS_JUDUL:
            pesan.append('judul {} karakter: {}'.format(len(L['judul']), L['judul'][:60]))
        if not MIN_DESK <= len(L['deskripsi']) <= MAKS_DESK:
            pesan.append('deskripsi {} karakter: {}'.format(len(L['deskripsi']), L['judul'][:60]))
        for d in L['desain']:
            if len(d['sku']) > MAKS_VARIAN:
                pesan.append('nama varian (SKU) > {} karakter: {}'.format(MAKS_VARIAN, d['sku']))
        if len({d['sku'] for d in L['desain']}) != n:
            pesan.append('ada SKU kembar di listing: {}'.format(L['judul'][:60]))
        if len({d['sku'] for d in L['desain']}) != n:
            pesan.append('ada nama varian kembar di listing: {}'.format(L['judul'][:60]))
        if not L['utama'][0]:
            pesan.append('BELUM ADA FOTO: {}'.format(L['judul'][:70]))
        for id_attr, label in wajib.get(L['jenis'], {}).items():
            if id_attr not in cfg['jenis'][L['jenis']]['atribut']:
                pesan.append('atribut wajib belum diisi ({}): {}'.format(label, L['jenis']))
    return pesan


def peta_url_db():
    """{toko: {KUNCI: url}} dari database foto. Kosong kalau database belum ada."""
    if not os.path.exists(DB_PATH):
        return {}
    sys.path.insert(0, os.path.join(AKAR, 'tools'))
    import gudang
    db = gudang.buka(DB_PATH)
    try:
        peta = gudang.peta_url(db)
        tertunda = gudang.belum_terunggah(db)
        if tertunda:
            print('[info] {} foto sudah disalin tapi BELUM di-push ke GitHub -> '
                  'URL-nya belum dipakai di Excel'.format(tertunda))
        return peta
    finally:
        db.close()


def kumpulkan(cfg, data):
    """{nama_file_output: (path_template, [listing, ...])} untuk semua toko.

    URL foto diambil dari database (hasil perintah "unggah"). Kalau database
    belum ada, jatuh ke cara lama: memindai foto-upload/ + base_url.
    """
    dari_db = peta_url_db()
    if dari_db:
        n = sum(len(v) for v in dari_db.values())
        print('[info] URL foto diambil dari database: {} foto'.format(n))
        manifest = None
    else:
        manifest, _ = pindai_foto(cfg)
    if not cfg['foto'].get('base_url'):
        print('[info] foto.base_url belum diisi di config.json -> kolom foto dikosongkan')

    paket = OrderedDict()
    for toko in cfg['toko']:
        if not (cocok(toko['nama'], 'toko') or cocok(toko['folder_foto'], 'toko')):
            continue
        per_template = OrderedDict()
        for jenis in data:
            if jenis not in cfg['jenis']:
                print('[info] jenis "{}" tidak ada di config.json, dilewati'.format(jenis))
                continue
            per_template.setdefault(cfg['jenis'][jenis]['template'], []).extend(
                susun_listing(cfg, data, toko, jenis, manifest, dari_db))
        for nama_tpl, listings in per_template.items():
            berkas = '{} - {}.xlsx'.format(toko['nama'], nama_tpl.title())
            paket[berkas] = (cfg['template'][nama_tpl], listings)
    return paket


def perintah_cek(cfg, data, diam=False):
    for t in info_template(cfg):
        if not t['ada']:
            print('[cek] ! template {} tidak ada: {}'.format(t['kunci'], t['path']))
        elif t['kategori_hilang']:
            print('[cek] ! template {} tidak memuat kategori: {}'.format(
                t['kunci'], ', '.join(t['kategori_hilang'])))
        elif t['umur_hari'] and t['umur_hari'] > 30:
            print('[cek] ! template {} sudah berumur {} hari. Shopee sering menolak '
                  'template lama - unduh ulang kalau upload ditolak.'.format(
                      t['kunci'], int(t['umur_hari'])))
    paket = kumpulkan(cfg, data)
    wajib = {}
    for jenis, j in cfg['jenis'].items():
        wb = openpyxl.load_workbook(os.path.join(AKAR, cfg['template'][j['template']]))
        wajib[jenis] = atribut_wajib(wb, j['kategori'])
        wb.close()
    total = 0
    for berkas, (_, listings) in paket.items():
        pesan = periksa(cfg, listings, wajib)
        total += len(pesan)
        if not diam:
            print('[cek] {}: {} listing, {} peringatan'.format(berkas, len(listings), len(pesan)))
            for p in pesan:
                print('        - ' + p)
    if total == 0:
        print('[cek] tidak ada masalah')
    return total


def perintah_build(cfg, data, sub=None):
    paket = kumpulkan(cfg, data)
    tujuan = os.path.join(dir_keluaran(), sub) if sub else dir_keluaran()
    terkunci = []
    for berkas, (tpl, listings) in paket.items():
        try:
            n = tulis_excel(cfg, os.path.join(AKAR, tpl), os.path.join(tujuan, berkas), listings)
        except PermissionError:
            terkunci.append(berkas)
            print('[build] {:<58} DILEWATI - berkas sedang dibuka'.format(berkas))
            continue
        berfoto = sum(1 for L in listings if L['utama'][0])
        print('[build] {:<58} {} listing / {} baris / {} berfoto'.format(
            berkas, len(listings), n, berfoto))
    print('[build] hasil di: {}'.format(tujuan))
    if terkunci:
        print('   ! {} berkas tidak bisa ditimpa karena sedang dibuka di Excel.'.format(len(terkunci)))
        print('     Tutup dulu berkas berikut lalu jalankan "build" lagi:')
        for b in terkunci:
            print('       - ' + b)


def main():
    p = argparse.ArgumentParser(description='Pembuat file Shopee Mass Upload')
    p.add_argument('perintah', choices=['impor', 'perbarui', 'template', 'pasang-hook', 'deteksi', 'unggah', 'foto', 'url', 'cek', 'build', 'semua'])
    p.add_argument('sumber', nargs='?', help='untuk "impor": berkas ekspor SKU; untuk "unggah": folder produk')
    p.add_argument('--tanpa-push', action='store_true', help='unggah: siapkan saja, jangan push ke GitHub')
    p.add_argument('--pasang', action='store_true', help='perbarui: langsung pasang, jangan cek saja')
    p.add_argument('--toko', help='uji coba: batasi ke satu toko, mis. "toko1" atau "Hangs"')
    p.add_argument('--jenis', help='uji coba: batasi ke satu jenis, mis. "JIBBITZ"')
    p.add_argument('--seri', help='uji coba: batasi ke satu seri, mis. "CORTIS"')
    a = p.parse_args()
    SARING.update(toko=a.toko, jenis=a.jenis, seri=a.seri)
    if menyaring():
        print('[uji] saringan aktif: {} -> hasil ditulis ke folder "uji", berkas asli tidak diubah'
              .format({k: v for k, v in SARING.items() if v}))
    cfg = baca_config()

    if a.perintah == 'impor':
        if not a.sumber:
            sys.exit('Contoh: python tools/shopee_mass_upload.py impor "SKU.xlsx"')
        perintah_impor(cfg, a.sumber)
        return

    if a.perintah == 'perbarui':
        import perbarui as modul_perbarui
        if a.pasang:
            modul_perbarui.pasang(sys.modules[__name__])
        else:
            info = modul_perbarui.periksa(sys.modules[__name__])
            if not info['siap']:
                print(info['pesan'])
            elif not info['jumlah']:
                print('Sudah versi terbaru ({}).'.format(info['sini']))
            else:
                print('{} pembaruan tersedia: {} -> {}'.format(
                    info['jumlah'], info['sini'], info['jauh']))
                for b in info['commit']:
                    print('   ' + b)
                print('')
                print('Pasang dengan: python tools/shopee_mass_upload.py perbarui --pasang')
        return

    if a.perintah == 'pasang-hook':
        asal = os.path.join(AKAR, 'tools', 'pre-commit')
        tujuan = os.path.join(AKAR, '.git', 'hooks', 'pre-commit')
        if not os.path.isdir(os.path.dirname(tujuan)):
            sys.exit('Folder .git/hooks tidak ada — project ini bukan repo git.')
        shutil.copy2(asal, tujuan)
        os.chmod(tujuan, 0o755)
        print('[hook] terpasang di .git/hooks/pre-commit')
        print('[hook] commit yang menghapus foto dari foto-upload/ akan ditolak')
        return

    if a.perintah == 'template':
        if not a.sumber:
            for t in info_template(cfg):
                print('{:<20} {:<34} {} kategori · {} hari · dipakai {}'.format(
                    t['kunci'], t['path'], t['kategori'],
                    int(t['umur_hari']) if t['umur_hari'] is not None else '?',
                    ', '.join(t['dipakai'])))
            print('\nUntuk mengganti: python tools/shopee_mass_upload.py template "unduhan.xlsx"')
            return
        pasang_template(cfg, a.sumber)
        return

    if a.perintah == 'deteksi':
        if not a.sumber:
            sys.exit('Contoh: python tools/shopee_mass_upload.py deteksi "G:/My Drive/JIBBITZ/..."')
        import unggah as modul_unggah
        modul_unggah.lapor_deteksi(sys.modules[__name__], cfg, a.sumber)
        return

    if a.perintah == 'unggah':
        if not a.sumber:
            sys.exit('Contoh: python tools/shopee_mass_upload.py unggah '
                     '"G:/My Drive/JIBBITZ/PRODUK 00001 - 00050"')
        import unggah as modul_unggah
        modul_unggah.proses(sys.modules[__name__], cfg, a.sumber, push=not a.tanpa_push)
        return

    data = baca_sku()
    jml = sum(len(d) for s in data.values() for d in s.values())
    print('[input] {} SKU, {} jenis, {} seri'.format(
        jml, len(data), sum(len(s) for s in data.values())))
    if a.perintah in ('foto', 'semua'):
        perintah_foto(cfg, data)
    if a.perintah == 'url' or (a.perintah == 'semua' and cfg['foto'].get('base_url')):
        perintah_url(cfg, data)
    if a.perintah in ('cek', 'semua'):
        perintah_cek(cfg, data)
    if a.perintah in ('build', 'semua'):
        perintah_build(cfg, data)


if __name__ == '__main__':
    main()
